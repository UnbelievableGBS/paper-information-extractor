"""
Export service - handles all export formats
Clean separation of concerns
"""

import os
import re
from typing import Optional, Dict, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import logging

from ..models import Paper, JournalType


class ExportService:
    """Service for exporting paper data to various formats"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def export_to_excel(self, paper: Paper, filename: Optional[str] = None) -> str:
        """
        Export paper to Excel format
        
        Args:
            paper: Paper object to export
            filename: Optional custom filename
            
        Returns:
            str: Path to created Excel file
        """
        if not filename:
            filename = self._generate_filename(paper, 'xlsx')
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Paper Information"
            
            # Apply journal-specific styling
            self._apply_excel_styling(ws, paper.journal)
            
            # Add headers
            headers = ["Field", "Value"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                self._apply_header_style(cell, paper.journal)
            
            # Add data rows
            data_rows = self._prepare_excel_data(paper)
            
            for row_idx, (field, value) in enumerate(data_rows, 2):
                ws.cell(row=row_idx, column=1, value=field)
                cell = ws.cell(row=row_idx, column=2, value=value)
                self._apply_data_cell_style(cell, field)
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 80
            
            # Save file
            wb.save(filename)
            self.logger.info(f"Excel file saved: {filename}")
            
            return filename
            
        except Exception as e:
            self.logger.error(f"Error creating Excel file: {e}")
            raise
    
    def export_to_json(self, paper: Paper, filename: Optional[str] = None) -> str:
        """Export paper to JSON format"""
        import json
        
        if not filename:
            filename = self._generate_filename(paper, 'json')
        
        try:
            data = paper.to_dict()
            
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            
            self.logger.info(f"JSON file saved: {filename}")
            return filename
            
        except Exception as e:
            self.logger.error(f"Error creating JSON file: {e}")
            raise
    
    def export_to_csv(self, paper: Paper, filename: Optional[str] = None) -> str:
        """Export paper to CSV format"""
        import csv
        
        if not filename:
            filename = self._generate_filename(paper, 'csv')
        
        try:
            data_rows = self._prepare_excel_data(paper)
            
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Field", "Value"])
                writer.writerows(data_rows)
            
            self.logger.info(f"CSV file saved: {filename}")
            return filename
            
        except Exception as e:
            self.logger.error(f"Error creating CSV file: {e}")
            raise
    
    def _generate_filename(self, paper: Paper, extension: str) -> str:
        """Generate filename from paper title"""
        title = paper.title or "paper"
        
        # Clean title for filename
        safe_title = re.sub(r'[^\w\s-]', '', title)
        safe_title = re.sub(r'[-\s]+', '_', safe_title)
        safe_title = safe_title[:50]  # Limit length
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{safe_title}_{timestamp}.{extension}"
    
    def _prepare_excel_data(self, paper: Paper) -> list:
        """Prepare data rows for Excel export"""
        data = [
            ("Title", paper.title),
            ("Journal", paper.journal.value.title()),
            ("DOI", paper.doi or "N/A"),
            ("Publication Date", paper.publication_date or "N/A"),
            ("URL", paper.url),
            ("Abstract", paper.abstract),
            ("Authors", self._format_authors_for_export(paper)),
            ("Author Count", str(paper.author_count)),
            ("Corresponding Authors", str(len(paper.corresponding_authors))),
            ("Volume", paper.volume or "N/A"),
            ("Issue", paper.issue or "N/A"),
            ("Pages", paper.pages or "N/A"),
            ("Extracted At", paper.extracted_at.isoformat() if paper.extracted_at else "N/A")
        ]
        
        return data
    
    def _format_authors_for_export(self, paper: Paper) -> str:
        """Format authors for export based on journal style"""
        if paper.journal == JournalType.NATURE:
            return paper.get_formatted_authors_string("nature")
        elif paper.journal == JournalType.APS:
            return paper.get_formatted_authors_string("aps")
        else:
            return paper.get_formatted_authors_string()
    
    def _apply_excel_styling(self, ws, journal_type: JournalType):
        """Apply journal-specific Excel styling"""
        # Journal-specific color schemes
        colors = {
            JournalType.NATURE: {"header": "1F4E79", "accent": "E7F3FF"},
            JournalType.SCIENCE: {"header": "C5504B", "accent": "FFF2F2"},
            JournalType.APS: {"header": "2F5233", "accent": "F0F8F0"}
        }
        
        # Store colors for use in styling methods
        ws._journal_colors = colors.get(journal_type, colors[JournalType.NATURE])
    
    def _apply_header_style(self, cell, journal_type: JournalType):
        """Apply header styling to cell"""
        colors = getattr(cell.parent, '_journal_colors', {"header": "1F4E79"})
        
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=colors["header"], end_color=colors["header"], fill_type="solid")
        cell.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _apply_data_cell_style(self, cell, field_name: str):
        """Apply data cell styling"""
        cell.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
        
        # Special styling for specific fields
        if field_name.lower() in ['abstract', 'authors']:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Increase row height for these fields
            cell.parent.row_dimensions[cell.row].height = 100
        else:
            cell.alignment = Alignment(vertical="center")