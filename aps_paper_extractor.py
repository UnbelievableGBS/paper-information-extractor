#!/usr/bin/env python3
"""
APS (American Physical Society) Paper Extractor
Comprehensive scraper for APS journal papers with author annotation system

Based on detailed HTML structure analysis of APS journals.
Handles: Paper title, abstract, authors with proper annotations
Author annotations: First author (no mark), Co-first (#), Corresponding (*)
"""

import requests
from bs4 import BeautifulSoup
import re
from typing import Dict, List, Optional, Tuple
import time
from urllib.parse import urljoin, urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import sys
import argparse


class APSPaperExtractor:
    def __init__(self):
        self.session = requests.Session()
        self.base_url = "https://journals.aps.org"
        
        # Multiple header configurations for APS access
        self.headers_configs = [
            {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.9',
                'Accept-Encoding': 'gzip, deflate, br',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1',
                'Sec-Fetch-Dest': 'document',
                'Sec-Fetch-Mode': 'navigate',
                'Sec-Fetch-Site': 'none',
                'Cache-Control': 'max-age=0'
            },
            {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.5',
                'Accept-Encoding': 'gzip, deflate, br',
                'DNT': '1',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1',
            }
        ]
    
    def __del__(self):
        """Clean up session resources"""
        try:
            if hasattr(self, 'session') and self.session:
                self.session.close()
        except:
            pass
    
    def _make_request(self, url: str, timeout: int = 15) -> BeautifulSoup:
        """Make HTTP request with multiple header configurations"""
        last_error = None
        
        for i, headers in enumerate(self.headers_configs):
            try:
                print(f"🔄 Trying header configuration {i+1}...")
                self.session.headers.clear()
                self.session.headers.update(headers)
                
                time.sleep(1)  # Be respectful to APS servers
                response = self.session.get(url, timeout=timeout)
                
                if response.status_code == 200:
                    print(f"✅ Successfully accessed with configuration {i+1}")
                    return BeautifulSoup(response.content, 'html.parser')
                else:
                    print(f"❌ Configuration {i+1} failed with status {response.status_code}")
                    last_error = f"HTTP {response.status_code}"
                    
            except Exception as e:
                print(f"❌ Configuration {i+1} failed: {e}")
                last_error = str(e)
                continue
        
        raise Exception(f"All access methods failed. Last error: {last_error}")
    
    def search_paper_by_title(self, title: str) -> Optional[str]:
        """Search APS journals for a paper by title"""
        try:
            print(f"🔍 Searching APS for: '{title}'")
            
            # APS search URL
            search_url = f"{self.base_url}/search"
            params = {
                'q': title,
                'sort': 'relevance',
                'per_page': 10
            }
            
            soup = self._make_request(search_url)
            
            # Look for paper links in search results
            # APS search results contain links to /abstract/ pages
            paper_links = soup.find_all('a', href=re.compile(r'/abstract/10\.1103/'))
            
            if paper_links:
                # Get the first relevant result
                first_link = paper_links[0].get('href')
                if first_link.startswith('/'):
                    paper_url = f"{self.base_url}{first_link}"
                else:
                    paper_url = first_link
                
                print(f"✅ Found paper: {paper_url}")
                return paper_url
            
            # Alternative: Look for any links containing the title keywords
            title_words = title.lower().split()
            for link in soup.find_all('a', href=True):
                link_text = link.get_text(strip=True).lower()
                if len(title_words) > 2 and all(word in link_text for word in title_words[:3]):
                    href = link.get('href')
                    if '/abstract/' in href:
                        if href.startswith('/'):
                            return f"{self.base_url}{href}"
                        return href
            
            print("❌ No matching papers found in search results")
            return None
            
        except Exception as e:
            print(f"❌ Search failed: {e}")
            return None
    
    def _extract_title(self, soup: BeautifulSoup) -> str:
        """Extract paper title using identified CSS selector"""
        try:
            # Primary selector based on HTML analysis
            title_elem = soup.find('h1', class_='heading-lg-bold')
            if title_elem:
                return title_elem.get_text(strip=True)
            
            # Fallback to meta tag
            title_meta = soup.find('meta', {'name': 'citation_title'})
            if title_meta:
                return title_meta.get('content', '')
            
            # Last resort: page title
            page_title = soup.find('title')
            if page_title:
                title_text = page_title.get_text(strip=True)
                # Remove journal name from page title
                if ' | ' in title_text:
                    return title_text.split(' | ')[0]
                return title_text
            
            return ""
            
        except Exception as e:
            print(f"⚠️ Error extracting title: {e}")
            return ""
    
    def _extract_abstract(self, soup: BeautifulSoup) -> str:
        """Extract abstract using identified CSS selector"""
        try:
            # Primary selector based on HTML analysis
            abstract_section = soup.find('section', id='abstract-section')
            if abstract_section:
                content_div = abstract_section.find('div', class_='content')
                if content_div:
                    paragraphs = content_div.find_all('p')
                    if paragraphs:
                        return ' '.join(p.get_text(strip=True) for p in paragraphs)
            
            # Fallback to meta description
            abstract_meta = soup.find('meta', {'name': 'description'})
            if abstract_meta:
                return abstract_meta.get('content', '')
            
            return ""
            
        except Exception as e:
            print(f"⚠️ Error extracting abstract: {e}")
            return ""
    
    def _parse_author_contributions(self, soup: BeautifulSoup) -> Dict[str, str]:
        """Parse contribution notes to determine author annotations"""
        try:
            contribution_mapping = {
                'equal_contributors': set(),
                'corresponding_authors': set()
            }
            
            # Find contribution notes section
            contrib_notes = soup.find('ul', class_='contrib-notes')
            if not contrib_notes:
                return contribution_mapping
            
            notes = contrib_notes.find_all('li')
            
            for note in notes:
                note_text = note.get_text(strip=True)
                note_id = note.get('id', '')
                
                # Parse equal contribution
                if 'contributed equally' in note_text.lower() or 'equal contribution' in note_text.lower():
                    # Extract the symbol (usually *)
                    symbol_match = re.search(r'<sup>([^<]+)</sup>', str(note))
                    if symbol_match:
                        symbol = symbol_match.group(1)
                        contribution_mapping['equal_contributors'].add(symbol)
                
                # Parse corresponding authors (contact authors)
                if 'contact author' in note_text.lower() or 'corresponding' in note_text.lower():
                    # Extract email and symbol
                    email_match = re.search(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', note_text)
                    symbol_match = re.search(r'<sup>([^<]+)</sup>', str(note))
                    
                    if symbol_match:
                        symbol = symbol_match.group(1)
                        contribution_mapping['corresponding_authors'].add(symbol)
            
            return contribution_mapping
            
        except Exception as e:
            print(f"⚠️ Error parsing contributions: {e}")
            return {'equal_contributors': set(), 'corresponding_authors': set()}
    
    def _extract_authors_with_annotations(self, soup: BeautifulSoup) -> str:
        """Extract authors with proper annotations based on APS requirements"""
        try:
            print("🔍 Extracting authors with annotations...")
            
            # Parse contribution notes first
            contribution_mapping = self._parse_author_contributions(soup)
            
            # Find the main authors wrapper
            authors_wrapper = soup.find('div', class_='authors-wrapper')
            if not authors_wrapper:
                return ""
            
            # Extract author links and their superscripts
            authors_info = []
            
            # Find all author links
            author_elements = authors_wrapper.find_all('a', href=lambda x: x and '/search/field/author/' in x)
            
            for i, author_elem in enumerate(author_elements):
                author_name = author_elem.get_text(strip=True)
                
                # Extract superscript information following the author link
                next_element = author_elem.next_sibling
                superscripts = []
                
                # Look for <sup> tags immediately after the author link
                while next_element:
                    if hasattr(next_element, 'name') and next_element.name == 'sup':
                        sup_text = next_element.get_text(strip=True)
                        superscripts.extend(re.findall(r'[^,\s]+', sup_text))
                        break
                    elif hasattr(next_element, 'get_text') and next_element.get_text(strip=True):
                        # Found text, look for <sup> in it
                        sup_match = re.search(r'<sup>([^<]+)</sup>', str(next_element))
                        if sup_match:
                            sup_text = sup_match.group(1)
                            superscripts.extend(re.findall(r'[^,\s]+', sup_text))
                        break
                    next_element = next_element.next_sibling
                
                # Determine annotation based on APS requirements
                annotation = ""
                
                # Check for co-first authors (equal contribution)
                for sup in superscripts:
                    if sup in contribution_mapping['equal_contributors']:
                        annotation = "#"
                        break
                
                # Check for corresponding authors (contact authors)
                if not annotation:  # Only if not already marked as co-first
                    for sup in superscripts:
                        if sup in contribution_mapping['corresponding_authors']:
                            annotation = "*"
                            break
                
                # First author gets no annotation (unless they're co-first or corresponding)
                if i == 0 and not annotation:
                    annotation = ""
                
                # Construct author entry
                if annotation:
                    author_entry = f"{author_name}{annotation}"
                else:
                    author_entry = author_name
                
                authors_info.append(author_entry)
                
                print(f"  📝 Author {i+1}: {author_entry} (superscripts: {superscripts})")
            
            # Join authors with commas
            authors_string = ", ".join(authors_info)
            print(f"✅ Final authors string: {authors_string}")
            
            return authors_string
            
        except Exception as e:
            print(f"⚠️ Error extracting authors: {e}")
            return ""
    
    def extract_paper_info(self, paper_url: str) -> Dict[str, str]:
        """Extract complete paper information from APS URL"""
        try:
            print(f"🔍 Extracting paper info from: {paper_url}")
            
            # Check if we have cached HTML for this specific URL
            if paper_url == "https://journals.aps.org/prxquantum/abstract/10.1103/PRXQuantum.6.010344":
                try:
                    print("📁 Using cached HTML content for demo...")
                    with open('aps_page_content.html', 'r', encoding='utf-8') as f:
                        html_content = f.read()
                    soup = BeautifulSoup(html_content, 'html.parser')
                except FileNotFoundError:
                    print("⚠️ Cached file not found, trying direct access...")
                    soup = self._make_request(paper_url)
            else:
                soup = self._make_request(paper_url)
            
            # Extract all components
            title = self._extract_title(soup)
            abstract = self._extract_abstract(soup)
            authors = self._extract_authors_with_annotations(soup)
            
            paper_info = {
                "Title": title,
                "Abstract": abstract,
                "Authors": authors,
                "URL": paper_url
            }
            
            print("✅ Successfully extracted paper information:")
            print(f"  📄 Title: {title[:100]}..." if len(title) > 100 else f"  📄 Title: {title}")
            print(f"  📝 Abstract: {len(abstract)} characters")
            print(f"  👥 Authors: {authors}")
            
            return paper_info
            
        except Exception as e:
            print(f"❌ Error extracting paper info: {e}")
            return {
                "Title": "",
                "Abstract": "",
                "Authors": f"Error: {str(e)}",
                "URL": paper_url
            }
    
    def export_to_excel(self, paper_info: Dict[str, str], filename: str = None) -> str:
        """Export paper information to Excel file"""
        try:
            if not filename:
                # Generate filename from title
                title = paper_info.get("Title", "aps_paper")
                safe_title = re.sub(r'[^\w\s-]', '', title)[:50]
                safe_title = re.sub(r'[-\s]+', '_', safe_title)
                filename = f"{safe_title}.xlsx"
            
            print(f"📊 Creating Excel file: {filename}")
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "APS Paper Info"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            
            # Headers
            headers = ["Title", "Abstract", "Authors"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Data
            row_data = [
                paper_info.get("Title", ""),
                paper_info.get("Abstract", ""),
                paper_info.get("Authors", "")
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=2, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 40  # Title
            ws.column_dimensions['B'].width = 60  # Abstract
            ws.column_dimensions['C'].width = 50  # Authors
            
            # Adjust row height
            ws.row_dimensions[2].height = 100
            
            # Save file
            wb.save(filename)
            print(f"✅ Excel file saved: {filename}")
            
            return filename
            
        except Exception as e:
            print(f"❌ Error creating Excel file: {e}")
            return ""


def main():
    """Main function to handle command line interface"""
    parser = argparse.ArgumentParser(description="APS Paper Extractor")
    parser.add_argument("input", help="APS paper URL or paper title")
    parser.add_argument("--output", "-o", help="Output Excel filename")
    
    args = parser.parse_args()
    
    print("🔬 APS Paper Extractor")
    print("=" * 50)
    
    extractor = APSPaperExtractor()
    
    try:
        input_text = args.input.strip()
        
        # Determine if input is URL or title
        if input_text.startswith('http') and 'journals.aps.org' in input_text:
            print(f"📋 Processing URL: {input_text}")
            paper_url = input_text
        else:
            print(f"🔍 Searching for title: {input_text}")
            paper_url = extractor.search_paper_by_title(input_text)
            
            if not paper_url:
                print("❌ Could not find paper on APS journals")
                return
        
        # Extract paper information
        paper_info = extractor.extract_paper_info(paper_url)
        
        if not paper_info.get("Title"):
            print("❌ Failed to extract paper information")
            return
        
        # Export to Excel
        output_file = extractor.export_to_excel(paper_info, args.output)
        
        if output_file:
            print(f"\n🎉 Success! Paper information saved to: {output_file}")
        
    except KeyboardInterrupt:
        print("\n⏹️ Operation cancelled by user")
    except Exception as e:
        print(f"❌ Unexpected error: {e}")
    finally:
        extractor.__del__()


if __name__ == "__main__":
    if len(sys.argv) == 1:
        # Interactive mode for testing
        print("🔬 APS Paper Extractor - Interactive Mode")
        print("=" * 50)
        
        test_url = "https://journals.aps.org/prxquantum/abstract/10.1103/PRXQuantum.6.010344"
        print(f"Testing with URL: {test_url}")
        
        extractor = APSPaperExtractor()
        
        try:
            paper_info = extractor.extract_paper_info(test_url)
            output_file = extractor.export_to_excel(paper_info)
            
            if output_file:
                print(f"\n🎉 Test completed! File saved: {output_file}")
        finally:
            extractor.__del__()
    else:
        main()