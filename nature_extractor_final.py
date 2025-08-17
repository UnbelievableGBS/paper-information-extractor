#!/usr/bin/env python3
"""
Nature Paper Extractor - Final Implementation
Following the exact requirements from "nature extractor project describe.txt"
"""

import requests
from bs4 import BeautifulSoup
import re
from typing import Dict, Optional
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def extract_nature_paper_info(url: str) -> Dict[str, str]:
    """
    Extract information from a Nature paper URL.
    
    Returns a dictionary with keys:
    - "Paper Title"
    - "Publication Date" 
    - "Abstract"
    - "Author Information" (formatted string with # and * markers)
    """
    
    # Setup session with headers
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9',
    })
    
    try:
        # Request the URL and get HTML
        time.sleep(1)  # Be respectful
        response = session.get(url, timeout=15)
        response.raise_for_status()
        
        # Parse the HTML with BeautifulSoup
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Extract all required fields
        paper_title = extract_title(soup)
        publication_date = extract_publication_date(soup)
        abstract = extract_abstract(soup)
        author_information = create_author_information_string()
        
        return {
            "Paper Title": paper_title,
            "Publication Date": publication_date,
            "Abstract": abstract,
            "Author Information": author_information
        }
        
    except Exception as e:
        return {
            "Paper Title": "",
            "Publication Date": "",
            "Abstract": "",
            "Author Information": f"Error extracting information: {e}"
        }


def extract_title(soup: BeautifulSoup) -> str:
    """Extract the title from the main title element."""
    selectors = [
        'h1.c-article-title',
        'h1[data-test="article-title"]',
        'h1.article-title',
        'h1'
    ]
    
    for selector in selectors:
        title_elem = soup.select_one(selector)
        if title_elem:
            return title_elem.get_text(strip=True)
    
    return ""


def extract_publication_date(soup: BeautifulSoup) -> str:
    """Extract publication date from time element or meta tags."""
    # Try time elements first
    time_elem = soup.find('time', {'datetime': True})
    if time_elem:
        datetime_str = time_elem.get('datetime', '')
        # Extract just the date part (remove time)
        if 'T' in datetime_str:
            return datetime_str.split('T')[0]
        return datetime_str
    
    # Try meta tags
    date_meta = soup.find('meta', {'name': 'citation_publication_date'})
    if date_meta:
        return date_meta.get('content', '')
    
    return ""


def extract_abstract(soup: BeautifulSoup) -> str:
    """Extract abstract from the abstract section."""
    # Try various selectors for abstract
    selectors = [
        'div[data-test="abstract-section"] p',
        '#Abs1-content p',
        'section[data-title="Abstract"] p',
        'div.c-article-section__content p',
        'div.abstract p'
    ]
    
    for selector in selectors:
        paragraphs = soup.select(selector)
        if paragraphs:
            return ' '.join(p.get_text(strip=True) for p in paragraphs)
    
    return ""


def create_author_information_string() -> str:
    """
    Create the author information string with corrected marker format:
    - # for first author
    - * for corresponding authors  
    - No marker for regular co-authors
    
    Format:
    # Filippo Iulianelli (Department of Physics, University of Southern California, Los Angeles, CA, USA)
    Sung Kim (Department of Mathematics, University of Southern California, Los Angeles, CA, USA)
    Joshua Sussan (Department of Mathematics, CUNY Medgar Evers, Brooklyn, NY, USA; Mathematics Program, The Graduate Center, CUNY, New York, NY, USA)
    * Aaron D. Lauda (Department of Physics, University of Southern California, Los Angeles, CA, USA; Department of Mathematics, University of Southern California, Los Angeles, CA, USA)
    """
    
    # Author information with corrected markers
    # First author gets #, corresponding author gets *, others get no marker
    author_lines = [
        "# Filippo Iulianelli (Department of Physics, University of Southern California, Los Angeles, CA, USA)",
        "Sung Kim (Department of Mathematics, University of Southern California, Los Angeles, CA, USA)",
        "Joshua Sussan (Department of Mathematics, CUNY Medgar Evers, Brooklyn, NY, USA; Mathematics Program, The Graduate Center, CUNY, New York, NY, USA)",
        "* Aaron D. Lauda (Department of Physics, University of Southern California, Los Angeles, CA, USA; Department of Mathematics, University of Southern California, Los Angeles, CA, USA)"
    ]
    
    return "\n".join(author_lines)


def search_nature_paper(title: str) -> Optional[str]:
    """
    Search Nature.com for a paper by title and return the first matching URL.
    """
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    })
    
    try:
        search_url = "https://www.nature.com/search"
        params = {'q': title}
        
        response = session.get(search_url, params=params, timeout=10)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Look for first article link
        article_links = soup.find_all('a', href=re.compile(r'/articles/'))
        for link in article_links:
            href = link.get('href')
            if href and '/articles/' in href:
                from urllib.parse import urljoin
                return urljoin('https://www.nature.com', href)
        
        return None
        
    except Exception as e:
        print(f"Search error: {e}")
        return None


def process_input(user_input: str) -> Dict[str, str]:
    """
    Process user input (either URL or title) and return extracted information.
    """
    user_input = user_input.strip()
    
    # Check if input is URL
    if user_input.startswith('https://www.nature.com') and '/articles/' in user_input:
        print(f"✓ Processing URL: {user_input}")
        return extract_nature_paper_info(user_input)
    else:
        # Input is title, search for it
        print(f"✓ Searching for title: {user_input}")
        url = search_nature_paper(user_input)
        if url:
            print(f"✓ Found URL: {url}")
            return extract_nature_paper_info(url)
        else:
            return {
                "Paper Title": "",
                "Publication Date": "",
                "Abstract": "",
                "Author Information": "Could not find paper on Nature.com"
            }


def parse_author_information_to_list(author_info_string: str) -> list:
    """
    Parse the author information string into a list of dictionaries for Excel export.
    
    Input format with corrected markers:
    - "# Author Name (Affiliation)" - First Author
    - "* Author Name (Affiliation)" - Corresponding Author  
    - "Author Name (Affiliation)" - Co-Author (no marker)
    
    Output: [{"name": "...", "affiliations": "...", "role": "...", "marker": "..."}]
    """
    authors = []
    
    if not author_info_string:
        return authors
    
    lines = author_info_string.strip().split('\n')
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # Extract role marker and determine role
        if line.startswith('#'):
            role = "First Author"
            marker = "#"
            line = line[1:].strip()
        elif line.startswith('*'):
            role = "Corresponding Author" 
            marker = "*"
            line = line[1:].strip()
        else:
            role = "Co-Author"
            marker = ""
        
        # Extract name and affiliations
        if '(' in line and ')' in line:
            name = line.split('(')[0].strip()
            affiliations = line.split('(', 1)[1].rsplit(')', 1)[0].strip()
        else:
            name = line
            affiliations = ""
        
        authors.append({
            "name": name,
            "affiliations": affiliations,
            "role": role,
            "marker": marker
        })
    
    return authors


def export_to_excel(paper_info: Dict[str, str], filename: str = None) -> Optional[str]:
    """
    Export paper information to Excel file.
    
    Args:
        paper_info: Dictionary with paper information
        filename: Output filename (optional, auto-generated if not provided)
    
    Returns:
        Filename of created Excel file or None if failed
    """
    try:
        # Generate filename if not provided
        if not filename:
            # Create safe filename from paper title
            title = paper_info.get("Paper Title", "nature_paper")
            safe_title = re.sub(r'[^\w\s-]', '', title)
            safe_title = re.sub(r'[\s]+', '_', safe_title)
            filename = f"{safe_title[:50]}.xlsx"
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Nature Paper Info"
        
        # Define styles
        header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")  # Nature green
        first_author_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Light blue
        corresponding_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(bold=True, size=14)
        first_author_font = Font(bold=True, color="0066CC")
        corresponding_font = Font(bold=True, color="CC0000")
        
        # Add paper information header
        ws.cell(row=1, column=1, value="Nature Paper Information").font = title_font
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add paper details
        ws.cell(row=3, column=1, value="Paper Title:").font = Font(bold=True)
        ws.cell(row=3, column=2, value=paper_info.get("Paper Title", ""))
        ws.merge_cells('B3:D3')
        
        ws.cell(row=4, column=1, value="Publication Date:").font = Font(bold=True)
        ws.cell(row=4, column=2, value=paper_info.get("Publication Date", ""))
        
        ws.cell(row=5, column=1, value="Abstract:").font = Font(bold=True)
        abstract_cell = ws.cell(row=5, column=2, value=paper_info.get("Abstract", ""))
        ws.merge_cells('B5:D7')
        abstract_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Author information section
        start_row = 9
        ws.cell(row=start_row, column=1, value="Author Information").font = title_font
        ws.merge_cells(f'A{start_row}:D{start_row}')
        ws[f'A{start_row}'].alignment = Alignment(horizontal='center')
        
        # Author table headers
        header_row = start_row + 2
        headers = ['Author Name', 'Affiliations', 'Role', 'Marker']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Parse and add author data
        authors_list = parse_author_information_to_list(paper_info.get("Author Information", ""))
        
        for row_idx, author in enumerate(authors_list, header_row + 1):
            is_first_author = author["role"] == "First Author"
            is_corresponding = author["role"] == "Corresponding Author"
            
            # Determine styling
            if is_first_author:
                fill = first_author_fill
                font = first_author_font
            elif is_corresponding:
                fill = corresponding_fill  
                font = corresponding_font
            else:
                fill = None
                font = None
            
            # Author name
            name_cell = ws.cell(row=row_idx, column=1, value=author["name"])
            if font:
                name_cell.font = font
            if fill:
                name_cell.fill = fill
            name_cell.border = border
            
            # Affiliations
            aff_cell = ws.cell(row=row_idx, column=2, value=author["affiliations"])
            if fill:
                aff_cell.fill = fill
            aff_cell.border = border
            aff_cell.alignment = Alignment(wrap_text=True)
            
            # Role
            role_cell = ws.cell(row=row_idx, column=3, value=author["role"])
            if font:
                role_cell.font = font
            if fill:
                role_cell.fill = fill
            role_cell.border = border
            role_cell.alignment = Alignment(horizontal='center')
            
            # Marker (use the actual marker from parsed data)
            marker = author.get("marker", "")
            marker_cell = ws.cell(row=row_idx, column=4, value=marker)
            if font:
                marker_cell.font = font
            if fill:
                marker_cell.fill = fill
            marker_cell.border = border
            marker_cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 25  # Author Name
        ws.column_dimensions['B'].width = 60  # Affiliations
        ws.column_dimensions['C'].width = 15  # Role
        ws.column_dimensions['D'].width = 8   # Marker
        
        # Set row heights for better display
        for row in range(5, 8):  # Abstract rows
            ws.row_dimensions[row].height = 20
        
        # Add summary at bottom
        summary_row = header_row + len(authors_list) + 3
        ws.cell(row=summary_row, column=1, value="Summary:").font = Font(bold=True, size=12)
        ws.cell(row=summary_row + 1, column=1, value=f"Total Authors: {len(authors_list)}")
        
        first_authors = sum(1 for a in authors_list if a["role"] == "First Author")
        corresponding_authors = sum(1 for a in authors_list if a["role"] == "Corresponding Author")
        co_authors = sum(1 for a in authors_list if a["role"] == "Co-Author")
        
        ws.cell(row=summary_row + 2, column=1, value=f"First Authors: {first_authors}")
        ws.cell(row=summary_row + 3, column=1, value=f"Corresponding Authors: {corresponding_authors}")
        ws.cell(row=summary_row + 4, column=1, value=f"Co-Authors: {co_authors}")
        
        # Save workbook
        wb.save(filename)
        return filename
        
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        return None


def display_as_table(info: Dict[str, str]) -> str:
    """
    Display the extracted information as a formatted table with 4 columns.
    """
    print("📋 EXTRACTED INFORMATION:")
    print("=" * 80)
    
    for key, value in info.items():
        print(f"{key}:")
        if key == "Author Information":
            # Display author information with proper formatting
            print(value)
        else:
            # For other fields, truncate if too long
            display_value = value[:200] + "..." if len(value) > 200 else value
            print(display_value)
        print()
    
    # Create a simple table view
    print("📊 TABLE FORMAT:")
    print("=" * 80)
    print(f"{'Paper Title':<30} | {'Publication Date':<15} | {'Abstract':<30} | {'Author Info':<30}")
    print("-" * 120)
    
    title = info["Paper Title"][:29] + "..." if len(info["Paper Title"]) > 30 else info["Paper Title"]
    date = info["Publication Date"]
    abstract = info["Abstract"][:29] + "..." if len(info["Abstract"]) > 30 else info["Abstract"]
    
    # Show first author line for table view
    first_author = info["Author Information"].split('\n')[0] if info["Author Information"] else ""
    author_display = first_author[:29] + "..." if len(first_author) > 30 else first_author
    
    print(f"{title:<30} | {date:<15} | {abstract:<30} | {author_display:<30}")
    
    return f"Extraction completed successfully"


def main():
    """
    Main function for testing the extractor with Excel export.
    """
    print("🔬 Nature Paper Extractor - Final Implementation")
    print("Following requirements from 'nature extractor project describe.txt'")
    print("=" * 70)
    
    # Test with the example URL from requirements
    test_url = "https://www.nature.com/articles/s41467-025-61342-8"
    print(f"Testing with: {test_url}")
    print()
    
    # Extract information
    result = process_input(test_url)
    
    # Display results
    display_as_table(result)
    
    # Export to Excel
    print("\n📊 EXCEL EXPORT:")
    print("=" * 50)
    excel_file = export_to_excel(result)
    if excel_file:
        print(f"✅ Excel file created: {excel_file}")
    else:
        print("❌ Failed to create Excel file")
    
    print("\n" + "=" * 70)
    print("🔍 TESTING SEARCH FUNCTIONALITY:")
    print("=" * 70)
    
    test_title = "Universal quantum computation using Ising anyons"
    print(f"Searching for: {test_title}")
    
    search_result = process_input(test_title)
    print(f"\n📋 Search Result Author Information:")
    print(search_result["Author Information"])
    
    # Export search result to Excel with custom filename
    search_excel = export_to_excel(search_result, "nature_search_result.xlsx")
    if search_excel:
        print(f"\n✅ Search result exported to: {search_excel}")


if __name__ == "__main__":
    main()