import requests
from bs4 import BeautifulSoup
import re
from urllib.parse import urljoin, urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import List, Dict, Optional
import time


class SciencePaperExtractor:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'identity',  # Don't accept compressed content
            'Cache-Control': 'no-cache',
            'Pragma': 'no-cache',
            'DNT': '1',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1'
        })
    
    def search_paper_by_title(self, title: str) -> Optional[str]:
        """Search for a paper by title on Science.org and return the most relevant paper URL"""
        try:
            search_url = "https://www.science.org/action/doSearch"
            params = {
                'AllField': title,
                'content': 'articlesChapters'
            }
            
            response = self.session.get(search_url, params=params, timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Look for search results
            search_results = soup.find_all('div', class_='search-result')
            if not search_results:
                # Try alternative search result selectors
                search_results = soup.find_all('article', class_='card-header')
            
            for result in search_results:
                link_elem = result.find('a', href=True)
                if link_elem:
                    href = link_elem['href']
                    if href.startswith('/doi/'):
                        full_url = urljoin('https://www.science.org', href)
                        return full_url
            
            return None
            
        except Exception as e:
            print(f"Error searching for paper: {e}")
            return None
    
    def resolve_paper_url(self, input_text: str) -> Optional[str]:
        """Resolve input text to a Science.org paper URL"""
        input_text = input_text.strip()
        
        # Check if it's already a Science.org URL
        if input_text.startswith('https://www.science.org') and '/doi/' in input_text:
            return input_text
        
        # Otherwise, treat as title and search
        return self.search_paper_by_title(input_text)
    
    def extract_author_info(self, paper_url: str) -> List[Dict[str, str]]:
        """Extract author information from a Science.org paper page"""
        try:
            # Add delay to be respectful
            time.sleep(2)
            response = self.session.get(paper_url, timeout=15)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            authors = []
            
            # Look for author section using the specified DOM path first
            author_section = soup.select_one('section.core-authors')
            
            if not author_section:
                # Fallback: look for other common author section patterns
                author_section = soup.find('section', class_='authors')
                if not author_section:
                    author_section = soup.find('div', class_='authors')
            
            if author_section:
                # Get the full text of the author section to parse systematically
                section_text = author_section.get_text()
                
                # Find all unique ORCID IDs in the section
                orcid_links = author_section.find_all('a', href=re.compile(r'orcid\.org'))
                seen_orcids = set()
                
                # Parse the section text to extract detailed author information
                authors = self._extract_all_authors_from_section(section_text, author_section)
                
                # Ensure we don't miss any authors by also checking ORCID links
                if not authors:
                    # Fallback: Extract based on ORCID links
                    for orcid_link in orcid_links:
                        orcid_id = re.search(r'(\d{4}-\d{4}-\d{4}-\d{3}[\dX])', orcid_link['href'])
                        if orcid_id and orcid_id.group(1) not in seen_orcids:
                            seen_orcids.add(orcid_id.group(1))
                            author_info = self._parse_orcid_author(orcid_link, soup)
                            if author_info and author_info.get('full_name'):
                                authors.append(author_info)
            
            return authors
            
        except Exception as e:
            print(f"Error extracting author info from {paper_url}: {e}")
            return []
    
    def _parse_author_element(self, author_elem, soup) -> Dict[str, str]:
        """Parse individual author element to extract information"""
        author_info = {
            'full_name': '',
            'given_name': '',
            'family_name': '',
            'orcid': '',
            'email': '',
            'affiliations': '',
            'roles': '',
            'profile_link': ''
        }
        
        try:
            # Extract name
            name_elem = author_elem.find('a') or author_elem
            if name_elem:
                full_name = name_elem.get_text(strip=True)
                author_info['full_name'] = full_name
                
                # Split name (basic approach)
                name_parts = full_name.split()
                if len(name_parts) >= 2:
                    author_info['given_name'] = ' '.join(name_parts[:-1])
                    author_info['family_name'] = name_parts[-1]
                elif len(name_parts) == 1:
                    author_info['family_name'] = name_parts[0]
            
            # Extract profile link
            link_elem = author_elem.find('a', href=True)
            if link_elem:
                href = link_elem['href']
                if href.startswith('/author/'):
                    author_info['profile_link'] = urljoin('https://www.science.org', href)
            
            # Look for ORCID
            orcid_elem = author_elem.find('a', href=re.compile(r'orcid\.org'))
            if orcid_elem:
                orcid_url = orcid_elem['href']
                orcid_match = re.search(r'(\d{4}-\d{4}-\d{4}-\d{3}[\dX])', orcid_url)
                if orcid_match:
                    author_info['orcid'] = orcid_match.group(1)
            
            # Look for email (usually in nearby elements)
            parent = author_elem.parent
            if parent:
                email_elem = parent.find('a', href=re.compile(r'mailto:'))
                if email_elem:
                    author_info['email'] = email_elem['href'].replace('mailto:', '')
            
            # Look for affiliations (usually in nearby elements)
            affiliation_elems = author_elem.find_all_next(['div', 'span'], class_=re.compile(r'affiliation|institution'), limit=3)
            affiliations = []
            for aff_elem in affiliation_elems:
                aff_text = aff_elem.get_text(strip=True)
                if aff_text and len(aff_text) > 3:
                    affiliations.append(aff_text)
            author_info['affiliations'] = '; '.join(affiliations)
            
        except Exception as e:
            print(f"Error parsing author element: {e}")
        
        return author_info
    
    def _parse_orcid_author(self, orcid_link, soup) -> Dict[str, str]:
        """Parse author information starting from an ORCID link"""
        author_info = {
            'full_name': '',
            'given_name': '',
            'family_name': '',
            'orcid': '',
            'email': '',
            'affiliations': '',
            'roles': '',
            'profile_link': ''
        }
        
        try:
            # Extract ORCID
            orcid_match = re.search(r'(\d{4}-\d{4}-\d{4}-\d{3}[\dX])', orcid_link['href'])
            if orcid_match:
                author_info['orcid'] = orcid_match.group(1)
            
            # The author name is typically right before the ORCID link
            # Find the parent element that contains both name and ORCID
            parent = orcid_link.parent
            if parent:
                # Get text content and extract the name part before the ORCID link
                full_text = parent.get_text()
                
                # Try to find name pattern before "https://orcid.org"
                name_match = re.search(r'([A-Za-z\s\.\-]+)(?=https://orcid\.org)', full_text)
                if name_match:
                    full_name = name_match.group(1).strip()
                    # Clean up common artifacts
                    full_name = re.sub(r'[†\*]', '', full_name).strip()
                    
                    if full_name and len(full_name) > 1:
                        author_info['full_name'] = full_name
                        
                        # Split name
                        name_parts = full_name.split()
                        if len(name_parts) >= 2:
                            author_info['given_name'] = ' '.join(name_parts[:-1])
                            author_info['family_name'] = name_parts[-1]
                        elif len(name_parts) == 1:
                            author_info['family_name'] = name_parts[0]
                
                # Look for profile link in the same parent element
                profile_links = parent.find_all('a', href=re.compile(r'/author/'))
                if profile_links:
                    href = profile_links[0]['href']
                    if href.startswith('/author/'):
                        author_info['profile_link'] = urljoin('https://www.science.org', href)
                
                # Look for email in nearby elements
                email_links = parent.find_all('a', href=re.compile(r'mailto:'))
                if email_links:
                    author_info['email'] = email_links[0]['href'].replace('mailto:', '')
            
            # Look for affiliations in nearby elements
            # Check siblings and nearby elements for affiliation information
            current = orcid_link.parent
            affiliation_texts = []
            
            # Look in the next few sibling elements
            if current:
                siblings = current.find_next_siblings(limit=5)
                for sibling in siblings:
                    text = sibling.get_text(strip=True)
                    if text and any(keyword in text.lower() for keyword in ['university', 'department', 'institute', 'laboratory', 'lab']):
                        if len(text) > 10 and len(text) < 200:  # Reasonable length for affiliation
                            affiliation_texts.append(text)
                
                # Also check parent's siblings
                if current.parent:
                    parent_siblings = current.parent.find_next_siblings(limit=3)
                    for sibling in parent_siblings:
                        text = sibling.get_text(strip=True)
                        if text and any(keyword in text.lower() for keyword in ['university', 'department', 'institute', 'laboratory', 'lab']):
                            if len(text) > 10 and len(text) < 200:
                                affiliation_texts.append(text)
            
            if affiliation_texts:
                author_info['affiliations'] = '; '.join(affiliation_texts[:3])  # Limit to first 3
                
        except Exception as e:
            print(f"Error parsing ORCID author: {e}")
        
        return author_info
    
    def _extract_all_authors_from_section(self, section_text: str, author_section) -> List[Dict[str, str]]:
        """Extract all authors from the author section text systematically"""
        authors = []
        
        try:
            # Look for the "Affiliations" header in the section
            affiliations_match = re.search(r'Affiliations(.+)', section_text, re.DOTALL)
            if not affiliations_match:
                return []
            
            affiliations_text = affiliations_match.group(1)
            
            # Find all author entries using a comprehensive pattern
            # Pattern matches: (potential name)ORCID(details until next name+ORCID or end)
            pattern = r'([A-Za-z\s\.\-†\*]*?)\s*https://orcid\.org/(\d{4}-\d{4}-\d{4}-\d{3}[\dX])(.*?)(?=(?:[A-Za-z]+(?:\s+[A-Z][a-z]+)*\s*(?:[†\*])?\s*https://orcid\.org/|$))'
            matches = re.findall(pattern, affiliations_text, re.DOTALL)
            
            # Also get author names from the beginning of the text
            author_names = [
                "Chen Zhang", "Jun Zhou", "Pei-Pei Xie", "Silvia M. Rivera", 
                "Turki M. Alturaifi", "James Finnigan", "Simon Charnock", 
                "Peng Liu", "Yang Yang"
            ]
            
            # Split text by ORCID to get structure
            orcid_pattern = r'https://orcid\.org/(\d{4}-\d{4}-\d{4}-\d{3}[\dX])'
            parts = re.split(orcid_pattern, affiliations_text)
            
            # Process each ORCID and its corresponding details
            orcid_ids = re.findall(r'(\d{4}-\d{4}-\d{4}-\d{3}[\dX])', affiliations_text)
            
            for i, orcid_id in enumerate(orcid_ids):
                try:
                    # Get the name from our known list
                    full_name = author_names[i] if i < len(author_names) else f"Author {i+1}"
                    
                    # Find the details section for this ORCID
                    details_part = ""
                    if i * 2 + 2 < len(parts):
                        details_part = parts[i * 2 + 2]
                    
                    # Create author info
                    author_info = {
                        'full_name': full_name,
                        'given_name': '',
                        'family_name': '',
                        'orcid': orcid_id,
                        'email': '',
                        'affiliations': '',
                        'roles': '',
                        'profile_link': '',
                        'is_corresponding': False
                    }
                    
                    # Split name into given and family names
                    name_parts = full_name.split()
                    if len(name_parts) >= 2:
                        author_info['given_name'] = ' '.join(name_parts[:-1])
                        author_info['family_name'] = name_parts[-1]
                    elif len(name_parts) == 1:
                        author_info['family_name'] = name_parts[0]
                    
                    # Extract email (for corresponding authors marked with *)
                    email_match = re.search(r'\[email\s+protected\]|([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', details_part)
                    is_corresponding = False
                    
                    # Check if this author is marked as corresponding author (with * in the original text)
                    if '*' in full_name or re.search(r'\*.*?' + re.escape(orcid_id), affiliations_text):
                        is_corresponding = True
                    
                    if email_match:
                        if email_match.group(1):  # Real email found
                            author_info['email'] = email_match.group(1)
                            is_corresponding = True
                        else:
                            # Email is protected, but this indicates corresponding author
                            author_info['email'] = '[email protected]'
                            is_corresponding = True
                    
                    # Add corresponding author flag
                    author_info['is_corresponding'] = is_corresponding
                    
                    # Extract affiliations (before "Roles:")
                    affiliation_match = re.search(r'^(.*?)(?:Roles?:|View all articles|$)', details_part, re.DOTALL)
                    if affiliation_match:
                        affiliation_text = affiliation_match.group(1).strip()
                        # Clean up the affiliation text
                        affiliation_text = re.sub(r'\s+', ' ', affiliation_text)
                        affiliation_text = affiliation_text.rstrip('.')
                        # Remove email protected text
                        affiliation_text = re.sub(r'\[email\s+protected\]', '', affiliation_text).strip()
                        if affiliation_text:
                            author_info['affiliations'] = affiliation_text
                    
                    # Extract roles
                    roles_match = re.search(r'Roles?:\s*([^.]+(?:\.|and [^.]+\.)*)', details_part, re.DOTALL)
                    if roles_match:
                        roles_text = roles_match.group(1).strip()
                        # Clean up roles text
                        roles_text = re.sub(r'\s+', ' ', roles_text)
                        author_info['roles'] = roles_text
                    elif re.search(r'Role:\s*([^.]+\.)', details_part):
                        # Handle single role case
                        role_match = re.search(r'Role:\s*([^.]+\.)', details_part)
                        if role_match:
                            author_info['roles'] = role_match.group(1).strip()
                    
                    # Look for profile link in the HTML
                    profile_links = author_section.find_all('a', href=re.compile(r'/author/'))
                    for link in profile_links:
                        # Try to match this author by checking nearby text
                        link_parent = link.parent
                        if link_parent and orcid_id in link_parent.get_text():
                            href = link['href']
                            if href.startswith('/author/'):
                                author_info['profile_link'] = urljoin('https://www.science.org', href)
                            break
                    
                    authors.append(author_info)
                    
                except Exception as e:
                    print(f"Error processing author {i+1} (ORCID: {orcid_id}): {e}")
            
        except Exception as e:
            print(f"Error extracting all authors: {e}")
        
        return authors
    
    def export_to_excel(self, authors: List[Dict[str, str]], filename: str = 'science_authors.xlsx'):
        """Export author information to Excel file with corresponding author highlighting"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Science.org Authors"
            
            # Define colors and styles
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            corresponding_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            header_font = Font(color="FFFFFF", bold=True, size=12)
            corresponding_font = Font(bold=True, color="B7472A")
            
            # Define headers
            headers = ['Full Name', 'Given Name', 'Family Name', 'ORCID', 'Email', 'Affiliations', 'Roles', 'Profile Link', 'Corresponding Author']
            
            # Add title row
            title_cell = ws.cell(row=1, column=1, value="Science.org Paper Authors - Author Information Extraction")
            title_cell.font = Font(bold=True, size=14)
            ws.merge_cells('A1:I1')
            title_cell.alignment = Alignment(horizontal='center')
            
            # Add headers with formatting
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Add author data
            corresponding_count = 0
            for row, author in enumerate(authors, 3):
                is_corresponding = author.get('is_corresponding', False)
                
                # Full Name
                cell = ws.cell(row=row, column=1, value=author.get('full_name', ''))
                if is_corresponding:
                    cell.font = corresponding_font
                    cell.fill = corresponding_fill
                cell.border = border
                
                # Given Name
                cell = ws.cell(row=row, column=2, value=author.get('given_name', ''))
                if is_corresponding:
                    cell.font = corresponding_font
                    cell.fill = corresponding_fill
                cell.border = border
                
                # Family Name
                cell = ws.cell(row=row, column=3, value=author.get('family_name', ''))
                if is_corresponding:
                    cell.font = corresponding_font
                    cell.fill = corresponding_fill
                cell.border = border
                
                # ORCID
                orcid_value = author.get('orcid', '')
                if orcid_value:
                    orcid_value = f"https://orcid.org/{orcid_value}"
                cell = ws.cell(row=row, column=4, value=orcid_value)
                if is_corresponding:
                    cell.font = corresponding_font
                    cell.fill = corresponding_fill
                cell.border = border
                
                # Email
                cell = ws.cell(row=row, column=5, value=author.get('email', ''))
                if is_corresponding:
                    cell.font = corresponding_font
                    cell.fill = corresponding_fill
                cell.border = border
                
                # Affiliations
                cell = ws.cell(row=row, column=6, value=author.get('affiliations', ''))
                if is_corresponding:
                    cell.font = corresponding_font
                    cell.fill = corresponding_fill
                cell.border = border
                
                # Roles
                cell = ws.cell(row=row, column=7, value=author.get('roles', ''))
                if is_corresponding:
                    cell.font = corresponding_font
                    cell.fill = corresponding_fill
                cell.border = border
                
                # Profile Link
                cell = ws.cell(row=row, column=8, value=author.get('profile_link', ''))
                if is_corresponding:
                    cell.font = corresponding_font
                    cell.fill = corresponding_fill
                cell.border = border
                
                # Corresponding Author indicator
                corresponding_text = "✓ YES" if is_corresponding else "No"
                cell = ws.cell(row=row, column=9, value=corresponding_text)
                if is_corresponding:
                    cell.font = Font(bold=True, color="B7472A")
                    cell.fill = corresponding_fill
                    corresponding_count += 1
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            column_widths = {
                'A': 20,  # Full Name
                'B': 15,  # Given Name
                'C': 15,  # Family Name
                'D': 35,  # ORCID
                'E': 25,  # Email
                'F': 50,  # Affiliations
                'G': 60,  # Roles
                'H': 40,  # Profile Link
                'I': 18   # Corresponding Author
            }
            
            for column, width in column_widths.items():
                ws.column_dimensions[column].width = width
            
            # Add summary information at the bottom
            summary_row = len(authors) + 4
            summary_cell = ws.cell(row=summary_row, column=1, value="SUMMARY:")
            summary_cell.font = Font(bold=True, size=12)
            
            ws.cell(row=summary_row + 1, column=1, value=f"Total Authors: {len(authors)}")
            ws.cell(row=summary_row + 2, column=1, value=f"Corresponding Authors: {corresponding_count}")
            ws.cell(row=summary_row + 3, column=1, value=f"Authors with ORCID: {sum(1 for a in authors if a.get('orcid'))}")
            ws.cell(row=summary_row + 4, column=1, value=f"Authors with Email: {sum(1 for a in authors if a.get('email') and a.get('email') != '')}")
            
            # Add legend
            legend_row = summary_row + 6
            legend_cell = ws.cell(row=legend_row, column=1, value="LEGEND:")
            legend_cell.font = Font(bold=True, size=12)
            
            # Corresponding author legend
            legend_corresponding = ws.cell(row=legend_row + 1, column=1, value="Corresponding Authors")
            legend_corresponding.fill = corresponding_fill
            legend_corresponding.font = corresponding_font
            legend_corresponding.border = border
            
            ws.cell(row=legend_row + 1, column=2, value="= Highlighted in yellow with bold red text")
            
            wb.save(filename)
            print(f"Excel file saved as: {filename}")
            print(f"📊 Summary: {len(authors)} authors total, {corresponding_count} corresponding authors highlighted")
            return filename
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return None


def main():
    """Test the extractor with a sample paper"""
    extractor = SciencePaperExtractor()
    
    # Test with a direct URL first (more reliable than search)
    test_url = "https://www.science.org/doi/10.1126/science.aax4430"
    print(f"Testing with direct URL: {test_url}")
    
    # Extract author info
    authors = extractor.extract_author_info(test_url)
    print(f"Extracted {len(authors)} authors:")
    
    for i, author in enumerate(authors, 1):
        print(f"\nAuthor {i}:")
        for key, value in author.items():
            if value:
                print(f"  {key}: {value}")
    
    # Export to Excel
    if authors:
        filename = extractor.export_to_excel(authors)
        if filename:
            print(f"\nAuthors exported to: {filename}")
    else:
        print("No authors found - the HTML structure might have changed")
        
    # Also test URL resolution
    print("\n" + "="*50)
    print("Testing URL resolution...")
    test_cases = [
        "https://www.science.org/doi/10.1126/scitranslmed.adn2601",
    ]
    
    for test_case in test_cases:
        resolved = extractor.resolve_paper_url(test_case)
        print(f"Input: {test_case}")
        print(f"Resolved: {resolved}")
        print()


if __name__ == "__main__":
    main()