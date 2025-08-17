#!/usr/bin/env python3
"""
Excel export utilities for multi-journal paper extractor
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import Dict, List, Optional
import re


def parse_author_information_to_list(author_info_string: str) -> List[Dict[str, str]]:
    """
    Parse author information string into structured list
    Handles: # First Author, * Corresponding Author, Co-Author (no marker)
    """
    authors = []
    
    if not author_info_string:
        return authors
    
    lines = author_info_string.strip().split('\n')
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # Extract role marker and determine role
        if line.startswith('#'):
            role = "First Author"
            marker = "#"
            line = line[1:].strip()
        elif line.startswith('*'):
            role = "Corresponding Author" 
            marker = "*"
            line = line[1:].strip()
        else:
            role = "Co-Author"
            marker = ""
        
        # Extract name and affiliations
        if '(' in line and ')' in line:
            name = line.split('(')[0].strip()
            affiliations = line.split('(', 1)[1].rsplit(')', 1)[0].strip()
        else:
            name = line
            affiliations = ""
        
        authors.append({
            "name": name,
            "affiliations": affiliations,
            "role": role,
            "marker": marker
        })
    
    return authors


def export_multi_journal_to_excel(paper_info: Dict[str, str], filename: str = None) -> Optional[str]:
    """
    Export paper information to Excel with multi-journal support
    """
    try:
        # Generate filename if not provided
        if not filename:
            title = paper_info.get("Paper Title", "paper")
            journal = paper_info.get("Journal", "unknown")
            safe_title = re.sub(r'[^\w\s-]', '', title)
            safe_title = re.sub(r'[\s]+', '_', safe_title)
            filename = f"{journal}_{safe_title[:40]}.xlsx"
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Paper Information"
        
        # Define journal-specific colors
        journal_colors = {
            "Nature": {"header": "2F5233", "accent": "E8F5E8"},
            "APS": {"header": "1E3A8A", "accent": "EBF4FF"},
            "Physical Review Letters": {"header": "1E3A8A", "accent": "EBF4FF"},
            "Physical Review A": {"header": "1E3A8A", "accent": "EBF4FF"},
            "Physical Review B": {"header": "1E3A8A", "accent": "EBF4FF"},
            "PRX Quantum": {"header": "7C3AED", "accent": "F3E8FF"},
        }
        
        journal = paper_info.get("Journal", "Unknown")
        colors = journal_colors.get(journal, journal_colors["Nature"])
        
        # Define styles
        header_fill = PatternFill(start_color=colors["header"], end_color=colors["header"], fill_type="solid")
        accent_fill = PatternFill(start_color=colors["accent"], end_color=colors["accent"], fill_type="solid")
        first_author_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        corresponding_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(bold=True, size=14)
        first_author_font = Font(bold=True, color="0066CC")
        corresponding_font = Font(bold=True, color="CC0000")
        
        # Add paper information header
        ws.cell(row=1, column=1, value=f"{journal} Paper Information").font = title_font
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add paper details
        ws.cell(row=3, column=1, value="Paper Title:").font = Font(bold=True)
        ws.cell(row=3, column=2, value=paper_info.get("Paper Title", ""))
        ws.merge_cells('B3:E3')
        
        ws.cell(row=4, column=1, value="Journal:").font = Font(bold=True)
        journal_cell = ws.cell(row=4, column=2, value=journal)
        journal_cell.fill = accent_fill
        journal_cell.font = Font(bold=True)
        
        ws.cell(row=5, column=1, value="Publication Date:").font = Font(bold=True)
        ws.cell(row=5, column=2, value=paper_info.get("Publication Date", ""))
        
        ws.cell(row=6, column=1, value="DOI:").font = Font(bold=True)
        ws.cell(row=6, column=2, value=paper_info.get("DOI", ""))
        ws.merge_cells('B6:E6')
        
        ws.cell(row=7, column=1, value="URL:").font = Font(bold=True)
        ws.cell(row=7, column=2, value=paper_info.get("URL", ""))
        ws.merge_cells('B7:E7')
        
        ws.cell(row=8, column=1, value="Abstract:").font = Font(bold=True)
        abstract_cell = ws.cell(row=8, column=2, value=paper_info.get("Abstract", ""))
        ws.merge_cells('B8:E10')
        abstract_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Author information section
        start_row = 12
        ws.cell(row=start_row, column=1, value="Author Information").font = title_font
        ws.merge_cells(f'A{start_row}:E{start_row}')
        ws[f'A{start_row}'].alignment = Alignment(horizontal='center')
        
        # Author table headers
        header_row = start_row + 2
        headers = ['Author Name', 'Affiliations', 'Role', 'Marker', 'Journal']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Parse and add author data
        authors_list = parse_author_information_to_list(paper_info.get("Author Information", ""))
        
        for row_idx, author in enumerate(authors_list, header_row + 1):
            is_first_author = author["role"] == "First Author"
            is_corresponding = author["role"] == "Corresponding Author"
            
            # Determine styling
            if is_first_author:
                fill = first_author_fill
                font = first_author_font
            elif is_corresponding:
                fill = corresponding_fill  
                font = corresponding_font
            else:
                fill = None
                font = None
            
            # Author name
            name_cell = ws.cell(row=row_idx, column=1, value=author["name"])
            if font:
                name_cell.font = font
            if fill:
                name_cell.fill = fill
            name_cell.border = border
            
            # Affiliations
            aff_cell = ws.cell(row=row_idx, column=2, value=author["affiliations"])
            if fill:
                aff_cell.fill = fill
            aff_cell.border = border
            aff_cell.alignment = Alignment(wrap_text=True)
            
            # Role
            role_cell = ws.cell(row=row_idx, column=3, value=author["role"])
            if font:
                role_cell.font = font
            if fill:
                role_cell.fill = fill
            role_cell.border = border
            role_cell.alignment = Alignment(horizontal='center')
            
            # Marker
            marker = author.get("marker", "")
            marker_cell = ws.cell(row=row_idx, column=4, value=marker)
            if font:
                marker_cell.font = font
            if fill:
                marker_cell.fill = fill
            marker_cell.border = border
            marker_cell.alignment = Alignment(horizontal='center')
            
            # Journal (for multi-paper exports)
            journal_cell = ws.cell(row=row_idx, column=5, value=journal)
            journal_cell.fill = accent_fill
            journal_cell.border = border
            journal_cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 25  # Author Name
        ws.column_dimensions['B'].width = 50  # Affiliations
        ws.column_dimensions['C'].width = 18  # Role
        ws.column_dimensions['D'].width = 8   # Marker
        ws.column_dimensions['E'].width = 15  # Journal
        
        # Set row heights for better display
        for row in range(8, 11):  # Abstract rows
            ws.row_dimensions[row].height = 25
        
        # Add summary at bottom
        summary_row = header_row + len(authors_list) + 3
        ws.cell(row=summary_row, column=1, value="Summary:").font = Font(bold=True, size=12)
        ws.cell(row=summary_row + 1, column=1, value=f"Journal: {journal}")
        ws.cell(row=summary_row + 2, column=1, value=f"Total Authors: {len(authors_list)}")
        
        first_authors = sum(1 for a in authors_list if a["role"] == "First Author")
        corresponding_authors = sum(1 for a in authors_list if a["role"] == "Corresponding Author")
        co_authors = sum(1 for a in authors_list if a["role"] == "Co-Author")
        
        ws.cell(row=summary_row + 3, column=1, value=f"First Authors: {first_authors}")
        ws.cell(row=summary_row + 4, column=1, value=f"Corresponding Authors: {corresponding_authors}")
        ws.cell(row=summary_row + 5, column=1, value=f"Co-Authors: {co_authors}")
        
        # Add legend
        legend_row = summary_row + 7
        ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True, size=12)
        
        # First author legend
        legend_first = ws.cell(row=legend_row + 1, column=1, value="First Authors (#)")
        legend_first.fill = first_author_fill
        legend_first.font = first_author_font
        legend_first.border = border
        
        # Corresponding author legend
        legend_corresponding = ws.cell(row=legend_row + 2, column=1, value="Corresponding Authors (*)")
        legend_corresponding.fill = corresponding_fill
        legend_corresponding.font = corresponding_font
        legend_corresponding.border = border
        
        # Journal legend
        legend_journal = ws.cell(row=legend_row + 3, column=1, value=f"{journal} Papers")
        legend_journal.fill = accent_fill
        legend_journal.font = Font(bold=True)
        legend_journal.border = border
        
        # Save workbook
        wb.save(filename)
        return filename
        
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        return None


def export_comparison_excel(papers: List[Dict[str, str]], filename: str = "journal_comparison.xlsx") -> Optional[str]:
    """
    Export multiple papers from different journals for comparison
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Journal Comparison"
        
        # Define styles
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        nature_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        aps_fill = PatternFill(start_color="EBF4FF", end_color="EBF4FF", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Add title
        ws.cell(row=1, column=1, value="Multi-Journal Paper Comparison").font = Font(bold=True, size=14)
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Paper Title', 'Journal', 'Publication Date', 'DOI', 'First Author', 'Author Count', 'URL']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Add paper data
        for row_idx, paper in enumerate(papers, 4):
            journal = paper.get("Journal", "")
            
            # Determine row color based on journal
            if "Nature" in journal:
                row_fill = nature_fill
            elif "APS" in journal or "Physical Review" in journal or "PRX" in journal:
                row_fill = aps_fill
            else:
                row_fill = None
            
            # Extract first author from author information
            author_info = paper.get("Author Information", "")
            first_author = ""
            if author_info:
                lines = author_info.split('\n')
                for line in lines:
                    if line.strip().startswith('#'):
                        first_author = line.strip()[1:].split('(')[0].strip()
                        break
            
            # Count authors
            author_count = len([line for line in author_info.split('\n') if line.strip()])
            
            # Data row
            data = [
                paper.get("Paper Title", "")[:50] + "..." if len(paper.get("Paper Title", "")) > 50 else paper.get("Paper Title", ""),
                journal,
                paper.get("Publication Date", ""),
                paper.get("DOI", ""),
                first_author,
                str(author_count),
                paper.get("URL", "")
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                if row_fill:
                    cell.fill = row_fill
                cell.border = border
                if col == 1:  # Title column
                    cell.alignment = Alignment(wrap_text=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 40  # Title
        ws.column_dimensions['B'].width = 20  # Journal
        ws.column_dimensions['C'].width = 15  # Date
        ws.column_dimensions['D'].width = 25  # DOI
        ws.column_dimensions['E'].width = 20  # First Author
        ws.column_dimensions['F'].width = 12  # Author Count
        ws.column_dimensions['G'].width = 30  # URL
        
        wb.save(filename)
        return filename
        
    except Exception as e:
        print(f"Error creating comparison Excel file: {e}")
        return None